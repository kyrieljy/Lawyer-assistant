import copy
import base64
import hashlib
import hmac
import json
import os
import re
import shutil
import sqlite3
import sys
import tempfile
import uuid
import secrets
import string
import urllib.error
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import keyring
except Exception:
    keyring = None

try:
    sys.stdin.reconfigure(encoding="utf-8")
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


CORE_FIELDS = [
    ("sign_date", "签约时间", "date"),
    ("case_category", "案件类别", "select"),
    ("cause", "案由/罪名", "text"),
    ("representation_stage", "代理阶段", "select"),
    ("client_name", "委托主体", "text"),
    ("client_role", "委托主体地位", "select"),
    ("opposing_party", "对方当事人", "text"),
    ("opposing_role", "对方主体地位", "select"),
    ("court", "管辖法院", "text"),
    ("claim_amount", "诉讼标的（元）", "money"),
    ("attorney_fee", "律师服务费（元）", "text"),
    ("related_case_numbers", "关联案号", "long_text"),
    ("judge_contact", "法官联系方式", "long_text"),
    ("todo_summary", "待办事项", "long_text"),
    ("time_points", "时间点", "long_text"),
    ("manual_progress", "工作进度", "long_text"),
    ("status", "状态", "select"),
    ("remarks", "备注", "long_text"),
]

CASE_COLUMNS = [field[0] for field in CORE_FIELDS]

DEFAULT_SETTINGS = {
    "feishu_webhook_url": "",
    "feishu_webhook_secret": "",
    "llm_provider": "",
    "llm_base_url": "",
    "llm_api_key": "",
    "llm_model_name": "",
    "vlm_base_url": "",
    "vlm_api_key": "",
    "vlm_model_name": "",
    "local_ocr_enabled": "true",
    "local_ocr_model_path": "",
    "local_ocr_language": "chi_sim+eng",
    "local_ocr_timeout_seconds": "60",
    "tesseract_path": "",
    "excel_template_path": "",
    "export_directory": "",
    "backup_directory": "",
    "backup_retention_count": "14",
    "reminder_hour": "9",
    "lawyer_name": "律所",
    "reminder_grace_days": "0",
}

SECRET_KEYS = {
    "feishu_webhook_secret",
    "llm_api_key",
    "vlm_api_key",
}

SERVICE_NAME = "lawyer-case-assistant"
SESSION_DAYS = 14
DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin666"
INVITE_CODE_SETTING = "registration_invite_code"
INVITE_CODE_ALPHABET = string.ascii_letters + string.digits

CASE_STATUS_OPTIONS = [
    "财产保全",
    "待排庭",
    "待开庭",
    "待判决",
    "待执行",
    "等待判决",
    "调节完成",
    "恢复执行",
    "诉前调解",
    "已在网上立案",
    "已撤诉",
    "结案",
    "执行终本",
]

CLOSED_STATUSES = {"已撤诉", "结案", "执行终本"}
ACTIVE_STATUSES = [status for status in CASE_STATUS_OPTIONS if status not in CLOSED_STATUSES]
DEFAULT_CASE_STATUS = "已在网上立案"

MOJIBAKE_REPAIR_SETTING = "data_encoding_repair_v1"

MOJIBAKE_MARKERS = (
    "\ufffd",
    "\u3125",
    "\ue000",
    "锛",
    "鍦",
    "寰嬪",
    "妗堜欢",
    "杩涘",
    "绛剧",
    "鏃堕",
    "棿",
    "缃",
    "瀵规",
    "柟",
    "褰撲",
    "簨浜",
    "绠¤",
    "緰",
    "璐",
    "鐘",
    "搴",
    "淇",
    "鎻",
    "鏈",
    "榛",
    "蹇",
    "澶",
)

DEFAULT_EXPORT_MAPPING = [
    ("A", "序号", "special", "sequence"),
    ("B", "签约时间", "fixed", "sign_date"),
    ("C", "案件类别", "fixed", "case_category"),
    ("D", "案由/罪名", "fixed", "cause"),
    ("E", "代理阶段", "fixed", "representation_stage"),
    ("F", "委托主体", "fixed", "client_name"),
    ("G", "主体地位", "fixed", "client_role"),
    ("H", "对方当事人", "fixed", "opposing_party"),
    ("I", "主体地位", "fixed", "opposing_role"),
    ("J", "管辖法院", "fixed", "court"),
    ("K", "诉讼\n标的（元）", "fixed", "claim_amount"),
    ("L", "律师服务费\n（元）", "fixed", "attorney_fee"),
    ("M", "关联案号", "fixed", "related_case_numbers"),
    ("N", "法官联系方式", "fixed", "judge_contact"),
    ("O", "待办事项、时间点", "special", "todo_deadlines"),
    ("P", "工作进度", "fixed", "manual_progress"),
    ("Q", "事件进度摘要", "special", "progress_events"),
    ("R", "状态", "fixed", "status"),
    ("S", "备注", "fixed", "remarks"),
]

BUSINESS_DATA_TABLES = [
    "settings",
    "case_fields",
    "cases",
    "case_custom_values",
    "events",
    "documents",
    "deadlines",
    "export_mappings",
    "ocr_results",
    "reminder_logs",
]

BUSINESS_DATA_DELETE_ORDER = [
    "reminder_logs",
    "ocr_results",
    "deadlines",
    "documents",
    "events",
    "case_custom_values",
    "cases",
    "export_mappings",
    "case_fields",
    "settings",
]


def now():
    return datetime.now().isoformat(timespec="seconds")


def data_dir():
    root = os.environ.get("LAWYER_ASSISTANT_DATA_DIR")
    if not root:
        root = str(Path(__file__).resolve().parents[1] / ".lawyer-case-assistant-data")
    path = Path(root)
    path.mkdir(parents=True, exist_ok=True)
    return path


def db_path():
    db_dir = data_dir() / "data"
    db_dir.mkdir(parents=True, exist_ok=True)
    return db_dir / "app.db"


def files_dir():
    path = data_dir() / "files"
    path.mkdir(parents=True, exist_ok=True)
    return path


def backups_dir():
    path = data_dir() / "backups"
    path.mkdir(parents=True, exist_ok=True)
    return path


def connect():
    conn = sqlite3.connect(db_path())
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def hash_password(password, salt=None):
    if not salt:
        salt = os.urandom(16).hex()
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 200_000).hex()
    return f"pbkdf2_sha256${salt}${digest}"


def verify_password(password, stored_hash):
    try:
        method, salt, digest = stored_hash.split("$", 2)
    except ValueError:
        return False
    if method != "pbkdf2_sha256":
        return False
    candidate = hash_password(password, salt).split("$", 2)[2]
    return hmac.compare_digest(candidate, digest)


def new_id(prefix):
    return f"{prefix}_{uuid.uuid4().hex[:16]}"


def generate_invite_code():
    return "".join(secrets.choice(INVITE_CODE_ALPHABET) for _ in range(6))


def read_payload():
    raw = sys.stdin.read()
    if not raw.strip():
        return {}
    return json.loads(raw)


def row_to_dict(row):
    return dict(row) if row is not None else None


def mojibake_score(text):
    if not isinstance(text, str) or not text:
        return 0
    score = 0
    for marker in MOJIBAKE_MARKERS:
        score += text.count(marker) * 3
    score += sum(1 for char in text if "\ue000" <= char <= "\uf8ff")
    return score


def has_cjk(text):
    return any("\u4e00" <= char <= "\u9fff" for char in text)


def should_accept_mojibake_repair(original, candidate, encoding):
    if candidate == original or not has_cjk(candidate):
        return False
    try:
        if candidate.encode("utf-8").decode(encoding) != original:
            return False
    except UnicodeError:
        return False
    original_score = mojibake_score(original)
    candidate_score = mojibake_score(candidate)
    if original_score > 0 and candidate_score < original_score:
        return True
    return len(candidate) < len(original) and candidate_score <= original_score


def repair_mojibake_text(text):
    if not isinstance(text, str) or not text:
        return text
    for encoding in ("gbk", "gb18030"):
        try:
            candidate = text.encode(encoding).decode("utf-8")
        except UnicodeError:
            continue
        if should_accept_mojibake_repair(text, candidate, encoding):
            return candidate
    return text


def repair_legacy_mojibake(conn):
    repaired = 0
    tables = [
        row["name"]
        for row in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        ).fetchall()
    ]
    for table in tables:
        columns = [
            row["name"]
            for row in conn.execute(f"PRAGMA table_info({table})").fetchall()
            if str(row["type"]).upper().startswith("TEXT")
        ]
        if not columns:
            continue
        rows = conn.execute(
            f"SELECT rowid, {', '.join(columns)} FROM {table}"
        ).fetchall()
        for row in rows:
            updates = {}
            for column in columns:
                value = row[column]
                fixed = repair_mojibake_text(value)
                if fixed != value:
                    updates[column] = fixed
            if updates:
                assignments = ", ".join([f"{column} = ?" for column in updates])
                conn.execute(
                    f"UPDATE {table} SET {assignments} WHERE rowid = ?",
                    [*updates.values(), row["rowid"]],
                )
                repaired += len(updates)

    conn.execute(
        "INSERT OR REPLACE INTO settings(key, value, updated_at) VALUES (?, ?, ?)",
        (MOJIBAKE_REPAIR_SETTING, f"repaired:{repaired}", now()),
    )
    conn.commit()
    return repaired


def init_schema(conn):
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            username TEXT NOT NULL UNIQUE,
            full_name TEXT NOT NULL,
            position TEXT NOT NULL DEFAULT '',
            password_hash TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS user_sessions (
            token TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            created_at TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            last_seen_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS business_logs (
            id TEXT PRIMARY KEY,
            user_id TEXT,
            username TEXT NOT NULL DEFAULT '',
            action TEXT NOT NULL,
            target_type TEXT NOT NULL DEFAULT '',
            target_id TEXT NOT NULL DEFAULT '',
            detail TEXT NOT NULL DEFAULT '',
            ip TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS case_fields (
            id TEXT PRIMARY KEY,
            field_key TEXT NOT NULL UNIQUE,
            label TEXT NOT NULL,
            field_type TEXT NOT NULL,
            builtin INTEGER NOT NULL DEFAULT 0,
            visible INTEGER NOT NULL DEFAULT 1,
            active INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL,
            options_json TEXT NOT NULL DEFAULT '[]',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS cases (
            id TEXT PRIMARY KEY,
            sign_date TEXT,
            case_category TEXT,
            cause TEXT,
            representation_stage TEXT,
            client_name TEXT,
            client_role TEXT,
            opposing_party TEXT,
            opposing_role TEXT,
            court TEXT,
            claim_amount TEXT,
            attorney_fee TEXT,
            related_case_numbers TEXT,
            judge_contact TEXT,
            todo_summary TEXT,
            time_points TEXT,
            manual_progress TEXT,
            status TEXT,
            remarks TEXT,
            ai_progress_summary TEXT,
            ai_progress_confirmed_at TEXT,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS case_custom_values (
            case_id TEXT NOT NULL,
            field_id TEXT NOT NULL,
            value TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL,
            PRIMARY KEY (case_id, field_id),
            FOREIGN KEY(case_id) REFERENCES cases(id) ON DELETE CASCADE,
            FOREIGN KEY(field_id) REFERENCES case_fields(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS events (
            id TEXT PRIMARY KEY,
            case_id TEXT NOT NULL,
            event_date TEXT NOT NULL,
            direction TEXT NOT NULL,
            counterparty_type TEXT,
            counterparty_name TEXT,
            summary TEXT NOT NULL,
            deadline_text TEXT,
            deadline_date TEXT,
            source TEXT,
            ocr_text TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(case_id) REFERENCES cases(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS documents (
            id TEXT PRIMARY KEY,
            case_id TEXT NOT NULL,
            event_id TEXT,
            original_path TEXT NOT NULL,
            stored_path TEXT NOT NULL,
            file_name TEXT NOT NULL,
            file_hash TEXT NOT NULL,
            ocr_status TEXT NOT NULL DEFAULT 'pending',
            ocr_text TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(case_id) REFERENCES cases(id) ON DELETE CASCADE,
            FOREIGN KEY(event_id) REFERENCES events(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS deadlines (
            id TEXT PRIMARY KEY,
            case_id TEXT NOT NULL,
            event_id TEXT,
            deadline_date TEXT NOT NULL,
            title TEXT NOT NULL,
            source TEXT NOT NULL DEFAULT 'manual',
            confirmed INTEGER NOT NULL DEFAULT 0,
            remind_20 INTEGER NOT NULL DEFAULT 1,
            remind_7 INTEGER NOT NULL DEFAULT 1,
            remind_1 INTEGER NOT NULL DEFAULT 1,
            last_notified_at TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(case_id) REFERENCES cases(id) ON DELETE CASCADE,
            FOREIGN KEY(event_id) REFERENCES events(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS export_mappings (
            column_key TEXT PRIMARY KEY,
            column_label TEXT NOT NULL,
            source_type TEXT NOT NULL,
            field_key TEXT NOT NULL,
            enabled INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS ocr_results (
            id TEXT PRIMARY KEY,
            document_id TEXT NOT NULL,
            case_id TEXT NOT NULL,
            event_id TEXT,
            engine TEXT NOT NULL,
            raw_text TEXT NOT NULL DEFAULT '',
            extracted_json TEXT NOT NULL DEFAULT '{}',
            status TEXT NOT NULL DEFAULT 'pending_review',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(document_id) REFERENCES documents(id) ON DELETE CASCADE,
            FOREIGN KEY(case_id) REFERENCES cases(id) ON DELETE CASCADE,
            FOREIGN KEY(event_id) REFERENCES events(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS reminder_logs (
            id TEXT PRIMARY KEY,
            deadline_id TEXT NOT NULL,
            reminder_days INTEGER NOT NULL,
            channel TEXT NOT NULL,
            sent_at TEXT NOT NULL,
            response_text TEXT,
            UNIQUE(deadline_id, reminder_days, channel),
            FOREIGN KEY(deadline_id) REFERENCES deadlines(id) ON DELETE CASCADE
        );
        """
    )
    conn.commit()


def seed_defaults(conn):
    timestamp = now()
    admin = conn.execute("SELECT id FROM users WHERE username = ?", (DEFAULT_ADMIN_USERNAME,)).fetchone()
    if not admin:
        conn.execute(
            """
            INSERT INTO users(id, username, full_name, position, password_hash, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?)
            """,
            (
                new_id("user"),
                DEFAULT_ADMIN_USERNAME,
                "管理员",
                "管理员",
                hash_password(DEFAULT_ADMIN_PASSWORD),
                timestamp,
                timestamp,
            ),
        )
    for key, value in DEFAULT_SETTINGS.items():
        conn.execute(
            "INSERT OR IGNORE INTO settings(key, value, updated_at) VALUES (?, ?, ?)",
            (key, value, timestamp),
        )
    invite_code = conn.execute(
        "SELECT value FROM settings WHERE key = ?",
        (INVITE_CODE_SETTING,),
    ).fetchone()
    if not invite_code:
        conn.execute(
            "INSERT INTO settings(key, value, updated_at) VALUES (?, ?, ?)",
            (INVITE_CODE_SETTING, generate_invite_code(), timestamp),
        )
    for index, (field_key, label, field_type) in enumerate(CORE_FIELDS, start=1):
        options_json = json.dumps(CASE_STATUS_OPTIONS, ensure_ascii=False) if field_key == "status" else "[]"
        conn.execute(
            """
            INSERT OR IGNORE INTO case_fields(
                id, field_key, label, field_type, builtin, visible, active, sort_order,
                options_json, created_at, updated_at
            ) VALUES (?, ?, ?, ?, 1, 1, 1, ?, '[]', ?, ?)
            """,
            (f"field_{field_key}", field_key, label, field_type, index, timestamp, timestamp),
        )
        if field_key == "status":
            conn.execute(
                "UPDATE case_fields SET field_type = 'select', options_json = ?, updated_at = ? WHERE field_key = 'status'",
                (options_json, timestamp),
            )
        if field_key == "manual_progress":
            conn.execute(
                """
                UPDATE case_fields
                SET label = ?, updated_at = ?
                WHERE field_key = 'manual_progress' AND builtin = 1 AND label IN ('手动工作进度', '工作进度')
                """,
                (label, timestamp),
            )
    for index, (col, label, source_type, field_key) in enumerate(DEFAULT_EXPORT_MAPPING, start=1):
        conn.execute(
            """
            INSERT OR IGNORE INTO export_mappings(
                column_key, column_label, source_type, field_key, enabled, sort_order, updated_at
            ) VALUES (?, ?, ?, ?, 1, ?, ?)
            """,
            (col, label, source_type, field_key, index, timestamp),
        )
        if col in {"P", "Q", "R", "S"}:
            conn.execute(
                """
                UPDATE export_mappings
                SET column_label = ?, source_type = ?, field_key = ?, enabled = 1, sort_order = ?, updated_at = ?
                WHERE column_key = ?
                """,
                (label, source_type, field_key, index, timestamp, col),
            )
    conn.execute(
        "UPDATE cases SET status = ? WHERE status IS NULL OR status = '' OR status IN ('在办', '进行中')",
        (DEFAULT_CASE_STATUS,),
    )
    conn.execute(
        "UPDATE cases SET status = '结案' WHERE status IN ('已结案', '归档')",
    )
    conn.commit()


def init(_payload):
    with connect() as conn:
        init_schema(conn)
        seed_defaults(conn)
        repair_legacy_mojibake(conn)
    return get_state({})


def get_settings(conn):
    rows = conn.execute("SELECT key, value FROM settings").fetchall()
    settings = dict(DEFAULT_SETTINGS)
    settings.update({row["key"]: row["value"] for row in rows})
    if keyring:
        for key in SECRET_KEYS:
            if settings.get(key) == "__keyring__":
                try:
                    settings[key] = keyring.get_password(SERVICE_NAME, key) or ""
                except Exception:
                    settings[key] = ""
    return settings


def get_fields(conn):
    return [
        row_to_dict(row)
        for row in conn.execute("SELECT * FROM case_fields ORDER BY sort_order ASC, created_at ASC").fetchall()
    ]


def get_export_mappings(conn):
    return [
        row_to_dict(row)
        for row in conn.execute("SELECT * FROM export_mappings ORDER BY sort_order ASC").fetchall()
    ]


def get_cases(conn):
    rows = conn.execute(
        "SELECT * FROM cases WHERE active = 1 ORDER BY updated_at DESC, created_at DESC"
    ).fetchall()
    cases = []
    for row in rows:
        item = row_to_dict(row)
        custom_rows = conn.execute(
            "SELECT field_id, value FROM case_custom_values WHERE case_id = ?",
            (item["id"],),
        ).fetchall()
        item["custom_values"] = {r["field_id"]: r["value"] for r in custom_rows}
        cases.append(item)
    return cases


def get_state(_payload):
    with connect() as conn:
        init_schema(conn)
        seed_defaults(conn)
        repair_legacy_mojibake(conn)
        settings = get_settings(conn)
        settings.pop(INVITE_CODE_SETTING, None)
        return {
            "ok": True,
            "dataDir": str(data_dir()),
            "dbPath": str(db_path()),
            "settings": settings,
            "fields": get_fields(conn),
            "cases": get_cases(conn),
            "events": [row_to_dict(row) for row in conn.execute("SELECT * FROM events ORDER BY event_date DESC, created_at DESC").fetchall()],
            "documents": [row_to_dict(row) for row in conn.execute("SELECT * FROM documents ORDER BY created_at DESC").fetchall()],
            "deadlines": [row_to_dict(row) for row in conn.execute("SELECT * FROM deadlines ORDER BY deadline_date ASC").fetchall()],
            "ocrResults": [row_to_dict(row) for row in conn.execute("SELECT * FROM ocr_results ORDER BY created_at DESC").fetchall()],
            "exportMappings": get_export_mappings(conn),
        }


def save_settings(payload):
    settings = payload.get("settings", {})
    timestamp = now()
    with connect() as conn:
        for key, value in settings.items():
            stored_value = "" if value is None else str(value)
            if key in SECRET_KEYS and keyring and stored_value:
                try:
                    keyring.set_password(SERVICE_NAME, key, stored_value)
                    stored_value = "__keyring__"
                except Exception:
                    pass
            conn.execute(
                """
                INSERT INTO settings(key, value, updated_at)
                VALUES (?, ?, ?)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
                """,
                (key, stored_value, timestamp),
            )
        conn.commit()
    return get_state({})


def normalize_key(label):
    raw = "".join(ch.lower() if ch.isalnum() else "_" for ch in label.strip())
    raw = "_".join(part for part in raw.split("_") if part)
    return raw[:48] or f"custom_{uuid.uuid4().hex[:8]}"


def save_field(payload):
    field = payload.get("field", {})
    timestamp = now()
    field_id = field.get("id") or new_id("field")
    label = (field.get("label") or "").strip()
    if not label:
        raise ValueError("字段名称不能为空")
    field_type = field.get("field_type") or field.get("fieldType") or "text"
    options = field.get("options_json")
    if options is None:
        options = json.dumps(field.get("options", []), ensure_ascii=False)
    with connect() as conn:
        existing = conn.execute("SELECT * FROM case_fields WHERE id = ?", (field_id,)).fetchone()
        if existing:
            conn.execute(
                """
                UPDATE case_fields
                SET label = ?, field_type = ?, visible = ?, active = ?, sort_order = ?,
                    options_json = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    label,
                    field_type,
                    1 if field.get("visible", True) else 0,
                    1 if field.get("active", True) else 0,
                    int(field.get("sort_order") or existing["sort_order"]),
                    options,
                    timestamp,
                    field_id,
                ),
            )
        else:
            max_order = conn.execute("SELECT COALESCE(MAX(sort_order), 0) FROM case_fields").fetchone()[0]
            field_key = field.get("field_key") or normalize_key(label)
            original = field_key
            counter = 2
            while conn.execute("SELECT 1 FROM case_fields WHERE field_key = ?", (field_key,)).fetchone():
                field_key = f"{original}_{counter}"
                counter += 1
            conn.execute(
                """
                INSERT INTO case_fields(
                    id, field_key, label, field_type, builtin, visible, active, sort_order,
                    options_json, created_at, updated_at
                ) VALUES (?, ?, ?, ?, 0, ?, 1, ?, ?, ?, ?)
                """,
                (
                    field_id,
                    field_key,
                    label,
                    field_type,
                    1 if field.get("visible", True) else 0,
                    int(field.get("sort_order") or max_order + 1),
                    options,
                    timestamp,
                    timestamp,
                ),
            )
        conn.commit()
    return get_state({})


def reorder_fields(payload):
    timestamp = now()
    with connect() as conn:
        for index, field_id in enumerate(payload.get("fieldIds", []), start=1):
            conn.execute(
                "UPDATE case_fields SET sort_order = ?, updated_at = ? WHERE id = ?",
                (index, timestamp, field_id),
            )
        conn.commit()
    return get_state({})


def save_case(payload):
    case_data = payload.get("case", {})
    custom_values = payload.get("customValues", {})
    timestamp = now()
    case_id = case_data.get("id") or new_id("case")
    values = {column: case_data.get(column, "") for column in CASE_COLUMNS}
    if not values.get("status"):
        values["status"] = DEFAULT_CASE_STATUS
    with connect() as conn:
        existing = conn.execute("SELECT id FROM cases WHERE id = ?", (case_id,)).fetchone()
        if existing:
            assignments = ", ".join([f"{column} = ?" for column in CASE_COLUMNS])
            conn.execute(
                f"UPDATE cases SET {assignments}, updated_at = ? WHERE id = ?",
                [values[column] for column in CASE_COLUMNS] + [timestamp, case_id],
            )
        else:
            columns = ", ".join(["id"] + CASE_COLUMNS + ["created_at", "updated_at"])
            placeholders = ", ".join(["?"] * (1 + len(CASE_COLUMNS) + 2))
            conn.execute(
                f"INSERT INTO cases({columns}) VALUES ({placeholders})",
                [case_id] + [values[column] for column in CASE_COLUMNS] + [timestamp, timestamp],
            )
        for field_id, value in custom_values.items():
            conn.execute(
                """
                INSERT INTO case_custom_values(case_id, field_id, value, updated_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(case_id, field_id) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
                """,
                (case_id, field_id, "" if value is None else str(value), timestamp),
            )
        conn.commit()
    return get_state({})


def save_event(payload):
    event = payload.get("event", {})
    timestamp = now()
    event_id = event.get("id") or new_id("event")
    case_id = event.get("case_id")
    if not case_id:
        raise ValueError("事件必须关联案件")
    fields = [
        "case_id", "event_date", "direction", "counterparty_type", "counterparty_name",
        "summary", "deadline_text", "deadline_date", "source", "ocr_text",
    ]
    values = {key: event.get(key, "") for key in fields}
    values["event_date"] = values["event_date"] or datetime.now().strftime("%Y-%m-%d")
    values["direction"] = values["direction"] or "收到"
    with connect() as conn:
        existing = conn.execute("SELECT id FROM events WHERE id = ?", (event_id,)).fetchone()
        if existing:
            assignments = ", ".join([f"{key} = ?" for key in fields])
            conn.execute(
                f"UPDATE events SET {assignments}, updated_at = ? WHERE id = ?",
                [values[key] for key in fields] + [timestamp, event_id],
            )
        else:
            columns = ", ".join(["id"] + fields + ["created_at", "updated_at"])
            placeholders = ", ".join(["?"] * (1 + len(fields) + 2))
            conn.execute(
                f"INSERT INTO events({columns}) VALUES ({placeholders})",
                [event_id] + [values[key] for key in fields] + [timestamp, timestamp],
            )
        title = values.get("deadline_text") or values.get("summary") or "待确认期限"
        event_deadlines = conn.execute(
            "SELECT id FROM deadlines WHERE event_id = ? ORDER BY created_at ASC",
            (event_id,),
        ).fetchall()
        primary_deadline_id = event_deadlines[0]["id"] if event_deadlines else ""
        duplicate_deadline_ids = [row["id"] for row in event_deadlines[1:]]
        if duplicate_deadline_ids:
            placeholders = ",".join(["?"] * len(duplicate_deadline_ids))
            conn.execute(f"DELETE FROM reminder_logs WHERE deadline_id IN ({placeholders})", duplicate_deadline_ids)
            conn.execute(f"DELETE FROM deadlines WHERE id IN ({placeholders})", duplicate_deadline_ids)
        if values.get("deadline_date"):
            if primary_deadline_id:
                conn.execute(
                    """
                    UPDATE deadlines
                    SET case_id = ?, deadline_date = ?, title = ?, source = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        case_id,
                        values["deadline_date"],
                        title,
                        values.get("source") or "manual",
                        timestamp,
                        primary_deadline_id,
                    ),
                )
            else:
                existing_deadline = conn.execute(
                    """
                    SELECT id FROM deadlines
                    WHERE case_id = ? AND deadline_date = ? AND title = ?
                    """,
                    (case_id, values["deadline_date"], title),
                ).fetchone()
                if existing_deadline:
                    conn.execute(
                        "UPDATE deadlines SET event_id = ?, updated_at = ? WHERE id = ?",
                        (event_id, timestamp, existing_deadline["id"]),
                    )
                else:
                    conn.execute(
                        """
                        INSERT INTO deadlines(
                            id, case_id, event_id, deadline_date, title, source, confirmed,
                            created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?)
                        """,
                        (
                            new_id("deadline"),
                            case_id,
                            event_id,
                            values["deadline_date"],
                            title,
                            values.get("source") or "manual",
                            timestamp,
                            timestamp,
                        ),
                    )
        elif primary_deadline_id:
            conn.execute("DELETE FROM reminder_logs WHERE deadline_id = ?", (primary_deadline_id,))
            conn.execute("DELETE FROM deadlines WHERE id = ?", (primary_deadline_id,))
        conn.commit()
    return get_state({})


def confirm_deadline(payload):
    deadline_id = payload.get("deadlineId")
    confirmed = 1 if payload.get("confirmed", True) else 0
    with connect() as conn:
        row = conn.execute("SELECT deadline_date FROM deadlines WHERE id = ?", (deadline_id,)).fetchone()
        if confirmed and row and not payload.get("confirmPast"):
            try:
                if date.fromisoformat(row["deadline_date"]) < date.today():
                    return {
                        "ok": False,
                        "needsPastDeadlineConfirmation": True,
                        "error": "该期限早于今天，属于历史待办，请确认是否继续提醒或取消提醒。",
                    }
            except Exception:
                pass
        conn.execute(
            "UPDATE deadlines SET confirmed = ?, updated_at = ? WHERE id = ?",
            (confirmed, now(), deadline_id),
        )
        conn.commit()
    return get_state({})


def cancel_deadline(payload):
    deadline_id = payload.get("deadlineId")
    with connect() as conn:
        conn.execute(
            """
            UPDATE deadlines
            SET confirmed = 0, remind_20 = 0, remind_7 = 0, remind_1 = 0, updated_at = ?
            WHERE id = ?
            """,
            (now(), deadline_id),
        )
        conn.commit()
    return get_state({})


def delete_deadline(payload):
    deadline_id = payload.get("deadlineId")
    with connect() as conn:
        conn.execute("DELETE FROM reminder_logs WHERE deadline_id = ?", (deadline_id,))
        conn.execute("DELETE FROM deadlines WHERE id = ?", (deadline_id,))
        conn.commit()
    return get_state({})


def hash_file(path):
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def upload_files(payload):
    case_id = payload.get("caseId")
    event_id = payload.get("eventId") or None
    paths = payload.get("paths", [])
    if not case_id:
        raise ValueError("上传文件必须先选择案件")
    timestamp = now()
    with connect() as conn:
        case_row = conn.execute("SELECT * FROM cases WHERE id = ?", (case_id,)).fetchone()
        if not case_row:
            raise ValueError("案件不存在")
        safe_case_name = (case_row["client_name"] or case_row["cause"] or case_id).replace(os.sep, "_")
        target_dir = files_dir() / safe_case_name / datetime.now().strftime("%Y%m%d")
        target_dir.mkdir(parents=True, exist_ok=True)
        for source in paths:
            source_path = Path(source)
            if not source_path.exists():
                continue
            file_hash = hash_file(source_path)
            target = target_dir / source_path.name
            counter = 2
            while target.exists():
                target = target_dir / f"{source_path.stem}_{counter}{source_path.suffix}"
                counter += 1
            shutil.copy2(source_path, target)
            conn.execute(
                """
                INSERT INTO documents(
                    id, case_id, event_id, original_path, stored_path, file_name,
                    file_hash, ocr_status, ocr_text, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, 'pending', '', ?)
                """,
                (
                    new_id("doc"),
                    case_id,
                    event_id,
                    str(source_path),
                    str(target),
                    target.name,
                    file_hash,
                    timestamp,
                ),
            )
        conn.commit()
    return get_state({})


def soft_delete_case(payload):
    case_id = payload.get("caseId")
    with connect() as conn:
        conn.execute("UPDATE cases SET active = 0, updated_at = ? WHERE id = ?", (now(), case_id))
        conn.commit()
    return get_state({})


def delete_event(payload):
    event_id = payload.get("eventId")
    with connect() as conn:
        deadline_ids = [
            row["id"]
            for row in conn.execute("SELECT id FROM deadlines WHERE event_id = ?", (event_id,)).fetchall()
        ]
        if deadline_ids:
            placeholders = ",".join(["?"] * len(deadline_ids))
            conn.execute(f"DELETE FROM reminder_logs WHERE deadline_id IN ({placeholders})", deadline_ids)
            conn.execute(f"DELETE FROM deadlines WHERE id IN ({placeholders})", deadline_ids)
        conn.execute("DELETE FROM events WHERE id = ?", (event_id,))
        conn.commit()
    return get_state({})


def delete_document(payload):
    document_id = payload.get("documentId")
    remove_file = bool(payload.get("removeFile", False))
    with connect() as conn:
        row = conn.execute("SELECT stored_path FROM documents WHERE id = ?", (document_id,)).fetchone()
        conn.execute("DELETE FROM documents WHERE id = ?", (document_id,))
        conn.commit()
    if remove_file and row and row["stored_path"]:
        path = Path(row["stored_path"])
        try:
            if path.exists() and files_dir() in path.resolve().parents:
                path.unlink()
        except Exception:
            pass
    return get_state({})


def save_export_mappings(payload):
    mappings = payload.get("mappings", [])
    timestamp = now()
    with connect() as conn:
        for index, mapping in enumerate(mappings, start=1):
            conn.execute(
                """
                INSERT INTO export_mappings(
                    column_key, column_label, source_type, field_key, enabled, sort_order, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(column_key) DO UPDATE SET
                    column_label = excluded.column_label,
                    source_type = excluded.source_type,
                    field_key = excluded.field_key,
                    enabled = excluded.enabled,
                    sort_order = excluded.sort_order,
                    updated_at = excluded.updated_at
                """,
                (
                    mapping["column_key"],
                    mapping.get("column_label", mapping["column_key"]),
                    mapping.get("source_type", "fixed"),
                    mapping.get("field_key", ""),
                    1 if mapping.get("enabled", True) else 0,
                    index,
                    timestamp,
                ),
            )
        conn.commit()
    return get_state({})


def case_scope_clause(scope):
    mode = scope.get("mode", "all")
    params = []
    clause = "active = 1"
    if mode == "active":
        placeholders = ",".join(["?"] * len(CLOSED_STATUSES))
        clause += f" AND status NOT IN ({placeholders})"
        params.extend(sorted(CLOSED_STATUSES))
    elif mode == "closed":
        placeholders = ",".join(["?"] * len(CLOSED_STATUSES))
        clause += f" AND status IN ({placeholders})"
        params.extend(sorted(CLOSED_STATUSES))
    elif mode == "status":
        statuses = [item for item in scope.get("statuses", []) if item]
        if statuses:
            placeholders = ",".join(["?"] * len(statuses))
            clause += f" AND status IN ({placeholders})"
            params.extend(statuses)
    return clause, params


def event_progress(conn, case_id):
    events = conn.execute(
        "SELECT * FROM events WHERE case_id = ? ORDER BY event_date ASC, created_at ASC",
        (case_id,),
    ).fetchall()
    docs = conn.execute(
        "SELECT event_id, file_name FROM documents WHERE case_id = ? ORDER BY created_at ASC",
        (case_id,),
    ).fetchall()
    docs_by_event = {}
    for doc in docs:
        docs_by_event.setdefault(doc["event_id"] or "", []).append(doc["file_name"])
    parts = []
    for event in events:
        actor = " ".join(
            item for item in [event["direction"], event["counterparty_type"], event["counterparty_name"]] if item
        )
        line = f"{event['event_date']} {actor} {event['summary']}".strip()
        files = docs_by_event.get(event["id"], [])
        if files:
            line += " 附件：" + "、".join(files)
        parts.append(line)
    return "\n".join(parts)


def clean_ai_progress_summary(text):
    text = (text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    text = re.sub(r"```(?:\w+)?", "", text)
    text = text.replace("```", "")
    lines = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        compact = re.sub(r"\s+", "", line)
        if re.fullmatch(r"\|?\s*:?-{2,}:?\s*(\|\s*:?-{2,}:?\s*)+\|?", line):
            continue
        if line.startswith("|") and line.endswith("|"):
            cells = [cell.strip() for cell in line.strip("|").split("|")]
            if all(not cell for cell in cells):
                continue
            header_words = {"日期", "时间", "事件", "事项", "内容", "进展", "工作进度", "说明", "状态"}
            if any(cell in header_words for cell in cells) and not any(re.search(r"\d{4}[-年/\.]\d{1,2}", cell) for cell in cells):
                continue
            line = " ".join(cell for cell in cells if cell)
        if any(
            phrase in compact
            for phrase in (
                "以下是根据您提供",
                "以下是根据你提供",
                "以下是案件事件流水",
                "以下为根据",
                "适用于Excel",
                "适合Excel",
                "未编造任何内容",
                "按时间顺序列出关键信息",
                "每行对应一个独立事件",
            )
        ):
            continue
        line = re.sub(r"^#{1,6}\s*", "", line)
        line = re.sub(r"^\s*(?:[-*•]|\d+[.)、])\s*", "", line)
        line = line.replace("**", "").replace("__", "").replace("`", "")
        if line:
            lines.append(line)
    return "\n".join(lines).strip()


def todo_deadline_text(conn, case_row):
    parts = []
    if case_row["todo_summary"]:
        parts.append(case_row["todo_summary"])
    if case_row["time_points"]:
        parts.append(case_row["time_points"])
    deadlines = conn.execute(
        """
        SELECT deadline_date, title FROM deadlines
        WHERE case_id = ? AND confirmed = 1
        ORDER BY deadline_date ASC
        """,
        (case_row["id"],),
    ).fetchall()
    for deadline in deadlines:
        parts.append(f"{deadline['deadline_date']} {deadline['title']}")
    return "\n".join(parts)


def load_optional(module_name):
    try:
        return __import__(module_name)
    except Exception:
        return None


def extract_pdf_text(path):
    pypdf = load_optional("pypdf")
    if pypdf:
        try:
            reader = pypdf.PdfReader(str(path))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            if text.strip():
                return text
        except Exception:
            pass
    fitz = load_optional("fitz")
    if fitz:
        try:
            doc = fitz.open(str(path))
            text = "\n".join(page.get_text() or "" for page in doc)
            if text.strip():
                return text
        except Exception:
            pass
    return ""


def rapidocr_image(path):
    try:
        from rapidocr_onnxruntime import RapidOCR
    except Exception:
        return ""
    engine = RapidOCR()
    result, _elapsed = engine(str(path))
    if not result:
        return ""
    return "\n".join(item[1] for item in result if len(item) >= 2 and item[1])


def tesseract_image(path, settings):
    command = settings.get("tesseract_path") or "tesseract"
    language = settings.get("local_ocr_language") or "chi_sim+eng"
    if language.strip().lower() in {"ch", "zh", "cn", "中文"}:
        language = "chi_sim+eng"
    languages = []
    for item in [language, "chi_sim+eng", "eng"]:
        if item and item not in languages:
            languages.append(item)
    for item in languages:
        try:
            import subprocess
            completed = subprocess.run(
                [command, str(path), "stdout", "-l", item],
                capture_output=True,
                text=True,
                timeout=int(settings.get("local_ocr_timeout_seconds") or 60),
                encoding="utf-8",
                errors="ignore",
            )
            if completed.returncode == 0 and completed.stdout.strip():
                return completed.stdout
        except Exception:
            return ""
    return ""


def render_pdf_pages(path, max_pages=5):
    fitz = load_optional("fitz")
    if not fitz:
        return []
    rendered = []
    doc = fitz.open(str(path))
    temp_dir = Path(tempfile.mkdtemp(prefix="lawyer-ocr-"))
    for index, page in enumerate(doc[:max_pages]):
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        output = temp_dir / f"page-{index + 1}.png"
        pix.save(str(output))
        rendered.append(output)
    return rendered


def local_ocr(path, settings):
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        text = extract_pdf_text(path)
        if text.strip():
            return text, "local_pdf_text"
        pages = render_pdf_pages(path)
        parts = []
        for page in pages:
            parts.append(rapidocr_image(page) or tesseract_image(page, settings))
        return "\n".join(part for part in parts if part.strip()), "local_pdf_ocr"
    if suffix in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}:
        return rapidocr_image(path) or tesseract_image(path, settings), "local_image_ocr"
    if suffix == ".txt":
        return path.read_text(encoding="utf-8", errors="ignore"), "local_text"
    return "", "unsupported"


def data_url(path):
    mime = "image/png"
    suffix = path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        mime = "image/jpeg"
    elif suffix == ".webp":
        mime = "image/webp"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def openai_chat(base_url, api_key, model, messages, temperature=0.1):
    endpoint = base_url.rstrip("/")
    if not endpoint.endswith("/chat/completions"):
        endpoint = endpoint.rstrip("/") + "/chat/completions"
    body = json.dumps(
        {"model": model, "messages": messages, "temperature": temperature},
        ensure_ascii=False,
    ).encode("utf-8")
    request = urllib.request.Request(
        endpoint,
        data=body,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return payload["choices"][0]["message"]["content"]


def vlm_ocr(path, settings):
    if not settings.get("vlm_base_url") or not settings.get("vlm_api_key") or not settings.get("vlm_model_name"):
        raise ValueError("未配置 VLM/vLLM OCR Base URL、API Key 或 Model Name")
    images = []
    if path.suffix.lower() == ".pdf":
        images = render_pdf_pages(path, max_pages=3)
    elif path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp"}:
        images = [path]
    else:
        text, _engine = local_ocr(path, settings)
        return text
    content = [
        {
            "type": "text",
            "text": "请识别这份法律文书/邮寄单/传票中的全部文字，并保留日期、案号、法院、联系人、电话、期限信息。只输出识别文字。",
        }
    ]
    for image in images:
        content.append({"type": "image_url", "image_url": {"url": data_url(image)}})
    return openai_chat(
        settings["vlm_base_url"],
        settings["vlm_api_key"],
        settings["vlm_model_name"],
        [{"role": "user", "content": content}],
    )


def compact_text(text):
    return re.sub(r"\s+", " ", text or "").strip()


def normalize_chinese_date(text):
    if not text:
        return ""
    match = re.search(r"(\d{4})\s*[年./-]\s*(\d{1,2})\s*[月./-]\s*(\d{1,2})\s*日?", text)
    if not match:
        return ""
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"


def normalize_datetime_text(text):
    text = compact_text(text)
    date_part = normalize_chinese_date(text)
    time_match = re.search(r"(\d{1,2})\s*[:：]\s*(\d{2})", text)
    if date_part and time_match:
        return f"{date_part} {int(time_match.group(1)):02d}:{time_match.group(2)}"
    return date_part or text


def chinese_number_value(text):
    digits = {
        "〇": 0, "零": 0, "一": 1, "二": 2, "三": 3, "四": 4,
        "五": 5, "六": 6, "七": 7, "八": 8, "九": 9,
    }
    if not text:
        return 0
    if text.isdigit():
        return int(text)
    if len(text) == 1:
        return digits.get(text, 0)
    if "十" in text:
        left, _, right = text.partition("十")
        return (digits.get(left, 1) if left else 1) * 10 + (digits.get(right, 0) if right else 0)
    value = 0
    for char in text:
        value = value * 10 + digits.get(char, 0)
    return value


def normalize_legal_date(text):
    date_text = compact_text(text)
    normal = normalize_chinese_date(date_text)
    if normal:
        return normal
    match = re.search(r"([二〇零一二三四五六七八九十\d]{4})年([一二三四五六七八九十\d]{1,3})月([一二三四五六七八九十\d]{1,3})日", date_text)
    if not match:
        return ""
    year_text, month_text, day_text = match.groups()
    year = int("".join(str(chinese_number_value(char)) for char in year_text))
    month = chinese_number_value(month_text)
    day = chinese_number_value(day_text)
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return ""


def add_years(date_text, years):
    try:
        base = date.fromisoformat(date_text)
        return base.replace(year=base.year + years).isoformat()
    except ValueError:
        return ""


def add_days(date_text, days):
    try:
        return (date.fromisoformat(date_text) + timedelta(days=days)).isoformat()
    except ValueError:
        return ""


def label_value(text, label, stop_labels):
    pattern = rf"{re.escape(label)}\s*(.*?)(?={'|'.join(re.escape(item) for item in stop_labels)}|$)"
    match = re.search(pattern, text, re.S)
    if not match:
        return ""
    return compact_text(match.group(1))


def extract_summons_fields(text):
    flat = compact_text(text)
    if "传票" not in flat and "应到时间" not in flat and "被传唤人" not in flat:
        return {}
    labels = [
        "案号", "案由", "被传唤人", "住所", "传唤事由", "应到时间", "应到处所",
        "注意事项", "联系地址", "联系电话", "书记员",
    ]
    case_match = re.search(r"[（(]\s*\d{4}\s*[）)]\s*[\u4e00-\u9fa5A-Za-z0-9\s]+?号", flat)
    appearance_time_raw = label_value(flat, "应到时间", labels)
    phone_match = re.search(r"(?:\d{3,4}-)?\d{7,11}", flat)
    clerk_match = re.search(r"书记员[:：]?\s*([\u4e00-\u9fa5]{2,5})", flat)
    court_match = re.search(r"[\u4e00-\u9fa5]{2,40}人民法院", flat)
    return {
        "document_type": "传票",
        "court": court_match.group(0) if court_match else "",
        "case_number": compact_text(case_match.group(0)) if case_match else "",
        "cause": label_value(flat, "案由", labels),
        "summoned_person": label_value(flat, "被传唤人", labels),
        "summons_reason": label_value(flat, "传唤事由", labels),
        "appearance_time": normalize_datetime_text(appearance_time_raw),
        "appearance_date": normalize_chinese_date(appearance_time_raw),
        "appearance_place": label_value(flat, "应到处所", labels),
        "contact_phone": phone_match.group(0) if phone_match else "",
        "clerk": clerk_match.group(1) if clerk_match else "",
    }


def extract_payment_notice_fields(text):
    flat = compact_text(text)
    if not any(key in flat for key in ["缴费通知书", "交费通知书", "交费须知", "应收金额", "交款单位"]):
        return {}
    legal_date = r"(?:[\d]{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日|[\d]{4}[./-]\d{1,2}[./-]\d{1,2}|[二〇零一二三四五六七八九十\d]{4}年[一二三四五六七八九十\d]{1,3}月[一二三四五六七八九十\d]{1,3}日)"
    print_date_match = re.search(rf"打印日期[:：]?\s*({legal_date})", flat)
    print_date = normalize_legal_date(print_date_match.group(1)) if print_date_match else ""
    amount_match = re.search(r"应收金额\s*([\d,]+(?:\.\d{1,2})?)", flat)
    if not amount_match:
        amount_match = re.search(r"合计\s*([\d,]+(?:\.\d{1,2})?)\s*元", flat)
    if not amount_match:
        amount_match = re.search(r"(?:案件受理费|诉讼费用|受理费)[^\d]{0,20}([\d,]+(?:\.\d{1,2})?)\s*元", flat)
    amount = amount_match.group(1).replace(",", "") if amount_match else ""
    deadline_date = add_days(print_date, 7) if print_date else ""
    if not print_date and not amount:
        return {}
    return {
        "document_type": "缴费通知书",
        "print_date": print_date,
        "amount": amount,
        "deadline_date": deadline_date,
        "deadline_text": f"缴费{amount}元诉讼费用" if amount else "缴纳诉讼费用",
    }


def extract_civil_ruling_fields(text):
    flat = compact_text(text)
    if "民事裁定书" not in flat:
        return {}
    legal_date = r"(?:[二〇零一二三四五六七八九十\d]{4}年[一二三四五六七八九十\d]{1,3}月[一二三四五六七八九十\d]{1,3}日|\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)"
    period_matches = re.findall(rf"期限为\s*({legal_date})\s*(?:至|到)\s*({legal_date})\s*止?", flat)
    period_end = normalize_legal_date(period_matches[-1][1]) if period_matches else ""
    date_matches = re.findall(rf"({legal_date})", flat)
    ruling_date = normalize_legal_date(date_matches[-1]) if date_matches else ""
    active_text = re.split(r"期限须知|以下空白", flat, maxsplit=1)[0]
    asset_specs = [
        (
            "银行账户",
            1,
            [
                r"冻结被申请人[^。；]{0,140}(银行存款|银行账户|账户|存款)",
                r"冻结[^。；]{0,80}账户\d{6,}",
                r"银行[^。；]{0,80}账户\d{6,}",
            ],
            period_end,
        ),
        (
            "动产",
            2,
            [
                r"(查封|扣押)被申请人[^。；]{0,160}(车辆|机动车|汽车|机器|设备|动产)",
                r"(查封|扣押)[^。；]{0,120}(车辆|机动车|汽车|机器设备|生产设备)",
            ],
            "",
        ),
        (
            "不动产",
            3,
            [
                r"(查封|冻结)被申请人[^。；]{0,180}(房产|房屋|不动产|土地|房地产|不动产权)",
                r"(查封|冻结)[^。；]{0,140}(房产|房屋|土地|房地产|不动产权)",
            ],
            "",
        ),
    ]
    asset_deadlines = []
    for asset_type, years, patterns, explicit_end in asset_specs:
        if any(re.search(pattern, active_text) for pattern in patterns):
            deadline_date = explicit_end or (add_years(ruling_date, years) if ruling_date else "")
            asset_deadlines.append(
                {
                    "asset_type": asset_type,
                    "preservation_years": years,
                    "deadline_date": deadline_date,
                    "deadline_text": f"{asset_type}保全期限届满前申请续封（{years}年）",
                }
            )
    if not asset_deadlines:
        asset_type, years = "财产保全", 1
        if any(key in flat for key in ["银行账户", "银行存款", "存款"]):
            asset_type, years = "银行账户", 1
        elif any(key in flat for key in ["不动产", "房产", "房屋", "土地"]):
            asset_type, years = "不动产", 3
        elif any(key in flat for key in ["动产", "车辆", "机器设备", "设备"]):
            asset_type, years = "动产", 2
        deadline_date = period_end or (add_years(ruling_date, years) if ruling_date else "")
        asset_deadlines.append(
            {
                "asset_type": asset_type,
                "preservation_years": years,
                "deadline_date": deadline_date,
                "deadline_text": f"{asset_type}保全期限届满前申请续封（{years}年）",
            }
        )
    if not ruling_date and "保全" not in flat and "查封" not in flat and "冻结" not in flat:
        return {}
    primary = asset_deadlines[0]
    return {
        "document_type": "民事裁定书",
        "ruling_date": ruling_date,
        "asset_type": primary["asset_type"],
        "preservation_years": primary["preservation_years"],
        "deadline_date": primary["deadline_date"],
        "deadline_text": primary["deadline_text"],
        "asset_deadlines": asset_deadlines,
    }


def extract_civil_judgment_fields(text):
    flat = compact_text(text)
    if "民事判决书" not in flat and "判决如下" not in flat:
        return {}
    received_match = re.search(r"(?:收到|送达|签收)[^\d二〇零一二三四五六七八九十]{0,10}([\d]{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日|[二〇零一二三四五六七八九十\d]{4}年[一二三四五六七八九十\d]{1,3}月[一二三四五六七八九十\d]{1,3}日)", flat)
    received_date = normalize_legal_date(received_match.group(1)) if received_match else date.today().isoformat()
    deadline_date = add_days(received_date, 15)
    return {
        "document_type": "民事判决书",
        "received_date": received_date,
        "deadline_date": deadline_date,
        "deadline_text": "收到民事判决书后15日内提起上诉，逾期将丧失上诉权",
    }


def extract_ocr_candidates(text):
    normalized = text or ""
    date_patterns = [
        r"\d{4}\s*[年./-]\s*\d{1,2}\s*[月./-]\s*\d{1,2}\s*日?",
        r"\d{1,2}\s*月\s*\d{1,2}\s*日",
    ]
    dates = []
    for pattern in date_patterns:
        dates.extend(compact_text(item) for item in re.findall(pattern, normalized))
    case_numbers = [
        compact_text(item)
        for item in re.findall(r"[（(]\s*\d{4}\s*[）)]\s*[\u4e00-\u9fa5A-Za-z0-9\s]+?号", normalized)
    ]
    phones = re.findall(r"(?:\d{3,4}-)?\d{7,11}", normalized)
    courts = re.findall(r"[\u4e00-\u9fa5]{2,40}(?:人民法院|仲裁委员会|法庭)", normalized)
    deadline_keywords = []
    for line in normalized.splitlines():
        if any(key in line for key in ["期限", "届满", "提交", "举证", "开庭", "应到时间", "应到处所", "答辩", "上诉"]):
            deadline_keywords.append(line.strip())
    result = {
        "dates": sorted(set(dates)),
        "case_numbers": sorted(set(case_numbers)),
        "phones": sorted(set(phones)),
        "courts": sorted(set(courts)),
        "deadline_lines": deadline_keywords[:20],
    }
    summons = extract_summons_fields(normalized)
    if summons:
        result["summons"] = summons
    payment_notice = extract_payment_notice_fields(normalized)
    if payment_notice:
        result["payment_notice"] = payment_notice
    civil_ruling = extract_civil_ruling_fields(normalized)
    if civil_ruling:
        result["civil_ruling"] = civil_ruling
    civil_judgment = extract_civil_judgment_fields(normalized)
    if civil_judgment:
        result["civil_judgment"] = civil_judgment
    return result


def events_from_ocr_result(ocr_row, event, source_tag="ocr"):
    if event:
        event["source"] = event.get("source") or source_tag
        return [event]
    extracted = json.loads(ocr_row["extracted_json"] or "{}")
    summons = extracted.get("summons") or {}
    if summons:
        reason = summons.get("summons_reason") or "传唤"
        appearance_time = summons.get("appearance_time") or ""
        place = summons.get("appearance_place") or ""
        summary_parts = [f"收到法院传票：{reason}"]
        if summons.get("case_number"):
            summary_parts.append(f"案号：{summons['case_number']}")
        if summons.get("cause"):
            summary_parts.append(f"案由：{summons['cause']}")
        if summons.get("summoned_person"):
            summary_parts.append(f"被传唤人：{summons['summoned_person']}")
        if appearance_time:
            summary_parts.append(f"应到时间：{appearance_time}")
        if place:
            summary_parts.append(f"应到处所：{place}")
        return [{
            "case_id": ocr_row["case_id"],
            "event_date": datetime.now().strftime("%Y-%m-%d"),
            "direction": "收到",
            "counterparty_type": "法院",
            "counterparty_name": summons.get("court") or "",
            "summary": "；".join(summary_parts),
            "deadline_text": place or reason,
            "deadline_date": summons.get("appearance_date") or "",
            "source": source_tag,
            "ocr_text": ocr_row["raw_text"],
        }]

    civil_ruling = extracted.get("civil_ruling") or {}
    if civil_ruling:
        items = civil_ruling.get("asset_deadlines") or [civil_ruling]
        events = []
        for item in items:
            asset_type = item.get("asset_type") or civil_ruling.get("asset_type") or "财产保全"
            events.append(
                {
                    "case_id": ocr_row["case_id"],
                    "event_date": datetime.now().strftime("%Y-%m-%d"),
                    "direction": "收到",
                    "counterparty_type": "法院",
                    "counterparty_name": "",
                    "summary": f"收到民事裁定书：{asset_type}保全续封提醒",
                    "deadline_text": item.get("deadline_text") or civil_ruling.get("deadline_text") or "财产保全期限届满前申请续封",
                    "deadline_date": item.get("deadline_date") or civil_ruling.get("deadline_date") or "",
                    "source": source_tag,
                    "ocr_text": ocr_row["raw_text"],
                }
            )
        return events

    templates = [
        ("payment_notice", "收到缴费通知书"),
        ("civil_judgment", "收到民事判决书"),
    ]
    for key, summary in templates:
        item = extracted.get(key) or {}
        if item:
            return [{
                "case_id": ocr_row["case_id"],
                "event_date": datetime.now().strftime("%Y-%m-%d"),
                "direction": "收到",
                "counterparty_type": "法院",
                "counterparty_name": "",
                "summary": summary,
                "deadline_text": item.get("deadline_text") or summary,
                "deadline_date": item.get("deadline_date") or "",
                "source": source_tag,
                "ocr_text": ocr_row["raw_text"],
            }]
    return []


def event_from_ocr_result(ocr_row, event):
    events = events_from_ocr_result(ocr_row, event)
    return events[0] if events else {}


def run_ocr(payload):
    document_id = payload.get("documentId")
    mode = payload.get("mode", "auto")
    with connect() as conn:
        settings = get_settings(conn)
        doc = conn.execute("SELECT * FROM documents WHERE id = ?", (document_id,)).fetchone()
        if not doc:
            raise ValueError("文件不存在")
        path = Path(doc["stored_path"])
        if not path.exists():
            raise ValueError("归档文件不存在")
        if mode == "vlm":
            text = vlm_ocr(path, settings)
            engine = "vlm"
        elif mode == "auto" and settings.get("vlm_api_key"):
            text = vlm_ocr(path, settings)
            engine = "vlm"
        else:
            text, engine = local_ocr(path, settings)
        candidates = extract_ocr_candidates(text or "")
        status = "pending_review" if text.strip() else "failed"
        timestamp = now()
        conn.execute(
            "UPDATE documents SET ocr_status = ?, ocr_text = ? WHERE id = ?",
            (status, text, document_id),
        )
        conn.execute(
            """
            INSERT INTO ocr_results(
                id, document_id, case_id, event_id, engine, raw_text, extracted_json,
                status, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                new_id("ocr"),
                document_id,
                doc["case_id"],
                doc["event_id"],
                engine,
                text or "",
                json.dumps(candidates, ensure_ascii=False),
                status,
                timestamp,
                timestamp,
            ),
        )
        conn.commit()
    return get_state({})


def create_event_from_ocr(payload):
    ocr_id = payload.get("ocrResultId")
    event = payload.get("event", {})
    with connect() as conn:
        ocr = conn.execute("SELECT * FROM ocr_results WHERE id = ?", (ocr_id,)).fetchone()
        if not ocr:
            raise ValueError("OCR 结果不存在")
        source_tag = f"ocr:{ocr_id}"
        events = events_from_ocr_result(ocr, event, source_tag)
        existing_event_id = ocr["event_id"] or ""
        stale_ids = [
            row["id"]
            for row in conn.execute(
                "SELECT id FROM events WHERE source = ? OR id = ?",
                (source_tag, existing_event_id),
            ).fetchall()
        ]
        if stale_ids:
            placeholders = ",".join(["?"] * len(stale_ids))
            deadline_ids = [
                row["id"]
                for row in conn.execute(
                    f"SELECT id FROM deadlines WHERE event_id IN ({placeholders})",
                    stale_ids,
                ).fetchall()
            ]
            if deadline_ids:
                deadline_placeholders = ",".join(["?"] * len(deadline_ids))
                conn.execute(f"DELETE FROM reminder_logs WHERE deadline_id IN ({deadline_placeholders})", deadline_ids)
                conn.execute(f"DELETE FROM deadlines WHERE id IN ({deadline_placeholders})", deadline_ids)
            conn.execute(f"DELETE FROM events WHERE id IN ({placeholders})", stale_ids)
            conn.commit()
    first_event_id = ""
    for item in events:
        item.setdefault("case_id", ocr["case_id"])
        item.setdefault("source", source_tag)
        item.setdefault("ocr_text", ocr["raw_text"])
        item["id"] = new_id("event")
        if not first_event_id:
            first_event_id = item["id"]
        save_event({"event": item})
    with connect() as conn:
        conn.execute(
            "UPDATE ocr_results SET event_id = ?, updated_at = ? WHERE id = ?",
            (first_event_id or None, now(), ocr_id),
        )
        conn.commit()
    return get_state({})


def generate_progress_summary(payload):
    case_id = payload.get("caseId")
    with connect() as conn:
        settings = get_settings(conn)
        case_row = conn.execute("SELECT * FROM cases WHERE id = ?", (case_id,)).fetchone()
        if not case_row:
            raise ValueError("案件不存在")
        progress = event_progress(conn, case_id) or case_row["manual_progress"] or ""
    if not settings.get("llm_base_url") or not settings.get("llm_api_key") or not settings.get("llm_model_name"):
        lines = [line for line in progress.splitlines() if line.strip()]
        summary = "\n".join(lines[-12:]) if lines else "暂无事件流水。"
        return {"ok": True, "summary": summary, "usedLlm": False}
    prompt = (
        "请把以下案件事件流水整理成 Excel 单元格可直接使用的工作进度纯文本。\n"
        "硬性要求：\n"
        "1. 直接输出整理后的进度内容，不要写“以下是”“根据您提供”“适用于Excel”等抬头或说明。\n"
        "2. 不要使用 Markdown，不要输出表格，不要使用 | 分隔线，不要使用代码块。\n"
        "3. 每个关键事件单独一行，格式建议为“日期 事项”。\n"
        "4. 按时间顺序，保留关键日期、法院/当事人动作、期限和当前状态。\n"
        "5. 不要编造事实；缺失的信息不要补写；总字数不超过800字。\n\n"
        "案件事件流水：\n"
        + progress
    )
    summary = openai_chat(
        settings["llm_base_url"],
        settings["llm_api_key"],
        settings["llm_model_name"],
        [
            {
                "role": "system",
                "content": "你是律师案件进度助理。你的输出会直接写入 Excel 单元格，只能输出纯文本进度内容。",
            },
            {"role": "user", "content": prompt},
        ],
    )
    return {"ok": True, "summary": clean_ai_progress_summary(summary), "usedLlm": True}


def save_progress_summary(payload):
    case_id = payload.get("caseId")
    summary = clean_ai_progress_summary(payload.get("summary", ""))
    with connect() as conn:
        conn.execute(
            "UPDATE cases SET ai_progress_summary = ?, ai_progress_confirmed_at = ?, updated_at = ? WHERE id = ?",
            (summary, now(), now(), case_id),
        )
        conn.commit()
    return get_state({})


def feishu_sign(secret, timestamp):
    string_to_sign = f"{timestamp}\n{secret}".encode("utf-8")
    digest = hmac.new(string_to_sign, b"", digestmod="sha256").digest()
    return base64.b64encode(digest).decode("utf-8")


def post_json(url, payload):
    request = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        return response.read().decode("utf-8", errors="ignore")


def send_feishu_text(settings, text):
    webhook = settings.get("feishu_webhook_url", "")
    if not webhook:
        raise ValueError("未配置飞书 Webhook URL")
    payload = {"msg_type": "text", "content": {"text": text}}
    secret = settings.get("feishu_webhook_secret", "")
    if secret:
        timestamp = str(int(datetime.now().timestamp()))
        payload["timestamp"] = timestamp
        payload["sign"] = feishu_sign(secret, timestamp)
    return post_json(webhook, payload)


def send_test_feishu(payload):
    with connect() as conn:
        settings = get_settings(conn)
    response = send_feishu_text(settings, payload.get("text") or "律师案件进度助手测试消息")
    return {"ok": True, "response": response}


def check_reminders(_payload):
    today = date.today()
    reminder_days = {20, 7, 1}
    sent = []
    skipped = []
    with connect() as conn:
        settings = get_settings(conn)
        rows = conn.execute(
            """
            SELECT d.*, c.client_name, c.cause, c.related_case_numbers
            FROM deadlines d
            JOIN cases c ON c.id = d.case_id
            WHERE d.confirmed = 1 AND c.active = 1
            ORDER BY d.deadline_date ASC
            """
        ).fetchall()
        for row in rows:
            try:
                deadline_day = date.fromisoformat(row["deadline_date"])
            except Exception:
                skipped.append({"deadlineId": row["id"], "reason": "invalid_date"})
                continue
            delta = (deadline_day - today).days
            if delta not in reminder_days:
                skipped.append({"deadlineId": row["id"], "reason": "not_due", "days": delta})
                continue
            flag_key = f"remind_{delta}"
            if flag_key in row.keys() and not row[flag_key]:
                skipped.append({"deadlineId": row["id"], "reason": "reminder_cancelled", "days": delta})
                continue
            exists = conn.execute(
                "SELECT 1 FROM reminder_logs WHERE deadline_id = ? AND reminder_days = ? AND channel = 'feishu'",
                (row["id"], delta),
            ).fetchone()
            if exists:
                skipped.append({"deadlineId": row["id"], "reason": "already_sent", "days": delta})
                continue
            title = row["client_name"] or row["cause"] or "未命名案件"
            text = (
                f"案件期限提醒（提前{delta}天）\n"
                f"案件：{title}\n"
                f"案号：{row['related_case_numbers'] or '未填写'}\n"
                f"期限：{row['deadline_date']}\n"
                f"事项：{row['title']}"
            )
            response = send_feishu_text(settings, text)
            conn.execute(
                """
                INSERT INTO reminder_logs(id, deadline_id, reminder_days, channel, sent_at, response_text)
                VALUES (?, ?, ?, 'feishu', ?, ?)
                """,
                (new_id("reminder"), row["id"], delta, now(), response),
            )
            sent.append({"deadlineId": row["id"], "days": delta, "title": row["title"]})
        conn.commit()
    return {
        "ok": True,
        "sent": sent,
        "skipped": skipped,
        "checked": len(rows),
        "note": "只会发送已确认且正好提前 20/7/1 天的期限；待确认、已取消、已发送过或不在提醒日的期限不会发送。",
    }


def apply_table_style(ws, row, column_count, header=False):
    thin = Side(style="thin", color="8C8C8C")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="E7ECEA") if header else PatternFill(fill_type=None)
    for col in range(1, column_count + 1):
        cell = ws.cell(row, col)
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Microsoft YaHei", bold=header, size=10)
        if header:
            cell.fill = fill


def create_default_workbook(settings, mappings):
    wb = Workbook()
    ws = wb.active
    ws.title = "在办案件"
    column_count = max(len(mappings), 1)
    last_column = get_column_letter(column_count)
    ws.merge_cells(f"A1:{last_column}1")
    ws["A1"] = f"{settings.get('lawyer_name') or '律所'}案件进度跟踪表"
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 55
    widths = [6.6, 16, 8, 14, 11, 16, 8, 18, 8, 16, 14, 18, 22, 24, 24, 90, 10, 24, 44]
    for idx in range(1, column_count + 1):
        width = widths[idx - 1] if idx <= len(widths) else 24
        ws.column_dimensions[get_column_letter(idx)].width = width
    for idx, mapping in enumerate(mappings, start=1):
        ws.cell(2, idx).value = mapping["column_label"]
    apply_table_style(ws, 2, column_count, header=True)
    apply_table_style(ws, 3, column_count, header=False)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_column}2"
    return wb


def prepare_workbook(template_path, settings, mappings):
    column_count = max(len(mappings), 1)
    last_column = get_column_letter(column_count)
    if template_path and Path(template_path).exists():
        wb = load_workbook(template_path)
        ws = wb.worksheets[0]
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row >= 3 or merged.max_row >= 3:
                ws.unmerge_cells(str(merged))
        style_row = [copy.copy(ws.cell(3, min(col, 18))._style) for col in range(1, column_count + 1)]
        if ws.max_row >= 3:
            ws.delete_rows(3, ws.max_row - 2)
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row == 1 and merged.max_row == 1:
                ws.unmerge_cells(str(merged))
        ws.merge_cells(f"A1:{last_column}1")
        ws["A1"] = f"{settings.get('lawyer_name') or '律所'}案件进度跟踪表"
        for idx, mapping in enumerate(mappings, start=1):
            ws.cell(2, idx).value = mapping["column_label"]
            if idx > 18:
                ws.cell(2, idx)._style = copy.copy(ws.cell(2, 18)._style)
                ws.column_dimensions[get_column_letter(idx)].width = 44
        return wb, ws, style_row
    wb = create_default_workbook(settings, mappings)
    ws = wb.worksheets[0]
    return wb, ws, [copy.copy(ws.cell(3, col)._style) for col in range(1, column_count + 1)]


def value_for_mapping(conn, case_row, custom_values, mapping, sequence):
    if not mapping.get("enabled", 1):
        return ""
    source_type = mapping["source_type"]
    field_key = mapping["field_key"]
    if source_type == "special":
        if field_key == "sequence":
            return sequence
        if field_key == "todo_deadlines":
            return todo_deadline_text(conn, case_row)
        if field_key == "progress_events":
            if case_row["ai_progress_summary"] and case_row["ai_progress_confirmed_at"]:
                return case_row["ai_progress_summary"]
            return event_progress(conn, case_row["id"])
        return ""
    if source_type == "custom":
        return custom_values.get(field_key, "")
    return case_row[field_key] if field_key in case_row.keys() else ""


def export_excel(payload):
    output_path = payload.get("outputPath")
    if not output_path:
        raise ValueError("请选择导出文件路径")
    scope = payload.get("scope", {"mode": "all"})
    with connect() as conn:
        settings = get_settings(conn)
        mappings = get_export_mappings(conn)
        column_count = max(len(mappings), 1)
        last_column = get_column_letter(column_count)
        clause, params = case_scope_clause(scope)
        cases = conn.execute(
            f"SELECT * FROM cases WHERE {clause} ORDER BY sign_date ASC, created_at ASC",
            params,
        ).fetchall()
        wb, ws, style_row = prepare_workbook(settings.get("excel_template_path", ""), settings, mappings)
        for row_idx, case_row in enumerate(cases, start=3):
            custom_rows = conn.execute(
                "SELECT field_id, value FROM case_custom_values WHERE case_id = ?",
                (case_row["id"],),
            ).fetchall()
            custom_values = {r["field_id"]: r["value"] for r in custom_rows}
            for col_idx, mapping in enumerate(mappings, start=1):
                cell = ws.cell(row_idx, col_idx)
                cell.value = value_for_mapping(conn, case_row, custom_values, mapping, row_idx - 2)
                if col_idx <= len(style_row):
                    cell._style = copy.copy(style_row[col_idx - 1])
                left_aligned = mapping["field_key"] in {"todo_deadlines", "progress_events", "manual_progress", "remarks"}
                cell.alignment = Alignment(
                    horizontal="left" if left_aligned else "center",
                    vertical="top" if left_aligned else "center",
                    wrap_text=True,
                )
            ws.row_dimensions[row_idx].height = 84
        ws.auto_filter.ref = f"A2:{last_column}{len(cases) + 2}" if cases else f"A2:{last_column}2"
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    return {"ok": True, "path": output_path, "count": len(cases)}


def make_backup(_payload):
    target_dir = backups_dir()
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    target = target_dir / f"app-{timestamp}.db"
    shutil.copy2(db_path(), target)
    return {"ok": True, "path": str(target)}


def public_user(row):
    if not row:
        return None
    data = row_to_dict(row)
    data.pop("password_hash", None)
    data["is_admin"] = data.get("username") == DEFAULT_ADMIN_USERNAME
    return data


def get_registration_invite_code():
    with connect() as conn:
        init_schema(conn)
        seed_defaults(conn)
        row = conn.execute(
            "SELECT value FROM settings WHERE key = ?",
            (INVITE_CODE_SETTING,),
        ).fetchone()
        return row["value"] if row else ""


def reset_registration_invite_code():
    code = generate_invite_code()
    timestamp = now()
    with connect() as conn:
        init_schema(conn)
        seed_defaults(conn)
        conn.execute(
            """
            INSERT INTO settings(key, value, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
            """,
            (INVITE_CODE_SETTING, code, timestamp),
        )
        conn.commit()
    return code


def invite_code_matches(conn, invite_code):
    invite_code = (invite_code or "").strip()
    row = conn.execute(
        "SELECT value FROM settings WHERE key = ?",
        (INVITE_CODE_SETTING,),
    ).fetchone()
    stored_code = row["value"] if row else ""
    return bool(invite_code and stored_code and hmac.compare_digest(invite_code, stored_code))


def register_user(username, full_name, position, password, invite_code=""):
    username = (username or "").strip()
    full_name = (full_name or "").strip()
    position = (position or "").strip()
    if not username:
        raise ValueError("用户名不能为空")
    if not full_name:
        raise ValueError("姓名不能为空")
    if len(password or "") <= 6:
        raise ValueError("密码必须大于6位")
    timestamp = now()
    with connect() as conn:
        init_schema(conn)
        seed_defaults(conn)
        if not invite_code_matches(conn, invite_code):
            raise ValueError("邀请码不正确")
        existing = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
        if existing:
            raise ValueError("用户名已存在")
        user_id = new_id("user")
        conn.execute(
            """
            INSERT INTO users(id, username, full_name, position, password_hash, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?)
            """,
            (user_id, username, full_name, position, hash_password(password), timestamp, timestamp),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        return public_user(row)


def authenticate_user(username, password):
    with connect() as conn:
        init_schema(conn)
        seed_defaults(conn)
        row = conn.execute("SELECT * FROM users WHERE username = ? AND active = 1", ((username or "").strip(),)).fetchone()
        if not row or not verify_password(password or "", row["password_hash"]):
            raise ValueError("用户名或密码错误")
        return public_user(row)


def create_session(user_id):
    token = base64.urlsafe_b64encode(os.urandom(32)).decode("ascii").rstrip("=")
    created_at = now()
    expires_at = (datetime.now() + timedelta(days=SESSION_DAYS)).isoformat(timespec="seconds")
    with connect() as conn:
        conn.execute(
            """
            INSERT INTO user_sessions(token, user_id, created_at, expires_at, last_seen_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (token, user_id, created_at, expires_at, created_at),
        )
        conn.commit()
    return token, expires_at


def get_user_by_session(token):
    if not token:
        return None
    timestamp = now()
    with connect() as conn:
        init_schema(conn)
        seed_defaults(conn)
        row = conn.execute(
            """
            SELECT users.*
            FROM user_sessions
            JOIN users ON users.id = user_sessions.user_id
            WHERE user_sessions.token = ? AND user_sessions.expires_at >= ? AND users.active = 1
            """,
            (token, timestamp),
        ).fetchone()
        if not row:
            return None
        conn.execute("UPDATE user_sessions SET last_seen_at = ? WHERE token = ?", (timestamp, token))
        conn.commit()
        return public_user(row)


def delete_session(token):
    if not token:
        return
    with connect() as conn:
        conn.execute("DELETE FROM user_sessions WHERE token = ?", (token,))
        conn.commit()


def log_business_action(user=None, action="", target_type="", target_id="", detail="", ip=""):
    timestamp = now()
    user = user or {}
    with connect() as conn:
        init_schema(conn)
        conn.execute(
            """
            INSERT INTO business_logs(id, user_id, username, action, target_type, target_id, detail, ip, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                new_id("log"),
                user.get("id") or "",
                user.get("username") or "",
                action,
                target_type or "",
                target_id or "",
                detail[:2000] if detail else "",
                ip or "",
                timestamp,
            ),
        )
        conn.commit()


def get_business_logs(limit=300):
    with connect() as conn:
        init_schema(conn)
        rows = conn.execute(
            """
            SELECT *
            FROM business_logs
            ORDER BY created_at DESC
            LIMIT ?
            """,
            (int(limit or 300),),
        ).fetchall()
        return [row_to_dict(row) for row in rows]


def list_users():
    with connect() as conn:
        init_schema(conn)
        seed_defaults(conn)
        rows = conn.execute(
            """
            SELECT *
            FROM users
            ORDER BY username = ? DESC, created_at ASC
            """,
            (DEFAULT_ADMIN_USERNAME,),
        ).fetchall()
        return [public_user(row) for row in rows]


def update_user(user_id, *, active=None, password=None):
    timestamp = now()
    with connect() as conn:
        init_schema(conn)
        seed_defaults(conn)
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            raise ValueError("账号不存在")
        updates = []
        params = []
        if active is not None:
            if row["username"] == DEFAULT_ADMIN_USERNAME and not active:
                raise ValueError("不能停用默认管理员账号")
            updates.append("active = ?")
            params.append(1 if active else 0)
        if password:
            if len(password) <= 6:
                raise ValueError("密码必须大于6位")
            updates.append("password_hash = ?")
            params.append(hash_password(password))
        if not updates:
            return public_user(row)
        updates.append("updated_at = ?")
        params.append(timestamp)
        params.append(user_id)
        conn.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
        conn.execute("DELETE FROM user_sessions WHERE user_id = ?", (user_id,))
        conn.commit()
        next_row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        return public_user(next_row)


def sqlite_table_exists(conn, table):
    return conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name = ?",
        (table,),
    ).fetchone() is not None


def sqlite_columns(conn, table):
    return [row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()]


def import_database_file(source_path):
    source = Path(source_path)
    if not source.exists():
        raise ValueError("上传的数据库文件不存在")
    backup = backups_dir() / f"before-import-{datetime.now().strftime('%Y%m%d-%H%M%S')}.db"
    if db_path().exists():
        shutil.copy2(db_path(), backup)
    source_conn = sqlite3.connect(source)
    source_conn.row_factory = sqlite3.Row
    try:
        if not sqlite_table_exists(source_conn, "cases") or not sqlite_table_exists(source_conn, "settings"):
            raise ValueError("上传文件不是有效的律师案件助手数据库")
        with connect() as conn:
            init_schema(conn)
            seed_defaults(conn)
            conn.execute("PRAGMA foreign_keys = OFF")
            for table in BUSINESS_DATA_DELETE_ORDER:
                conn.execute(f"DELETE FROM {table}")
            for table in BUSINESS_DATA_TABLES:
                if not sqlite_table_exists(source_conn, table):
                    continue
                dest_columns = sqlite_columns(conn, table)
                source_columns = sqlite_columns(source_conn, table)
                columns = [column for column in dest_columns if column in source_columns]
                if not columns:
                    continue
                column_sql = ", ".join(columns)
                placeholders = ", ".join(["?"] * len(columns))
                rows = source_conn.execute(f"SELECT {column_sql} FROM {table}").fetchall()
                for row in rows:
                    conn.execute(
                        f"INSERT OR REPLACE INTO {table}({column_sql}) VALUES ({placeholders})",
                        [row[column] for column in columns],
                    )
            conn.execute("PRAGMA foreign_keys = ON")
            seed_defaults(conn)
            repair_legacy_mojibake(conn)
            conn.commit()
        return {"ok": True, "backupPath": str(backup), "dbPath": str(db_path())}
    finally:
        source_conn.close()


COMMANDS = {
    "init": init,
    "getState": get_state,
    "saveSettings": save_settings,
    "saveField": save_field,
    "reorderFields": reorder_fields,
    "saveCase": save_case,
    "saveEvent": save_event,
    "confirmDeadline": confirm_deadline,
    "cancelDeadline": cancel_deadline,
    "deleteDeadline": delete_deadline,
    "uploadFiles": upload_files,
    "deleteCase": soft_delete_case,
    "deleteEvent": delete_event,
    "deleteDocument": delete_document,
    "runOcr": run_ocr,
    "createEventFromOcr": create_event_from_ocr,
    "generateProgressSummary": generate_progress_summary,
    "saveProgressSummary": save_progress_summary,
    "sendTestFeishu": send_test_feishu,
    "checkReminders": check_reminders,
    "saveExportMappings": save_export_mappings,
    "exportExcel": export_excel,
    "makeBackup": make_backup,
}


def main():
    command = sys.argv[1] if len(sys.argv) > 1 else "getState"
    if command not in COMMANDS:
        raise ValueError(f"Unknown command: {command}")
    result = COMMANDS[command](read_payload())
    print(json.dumps(result, ensure_ascii=False, default=str))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)
